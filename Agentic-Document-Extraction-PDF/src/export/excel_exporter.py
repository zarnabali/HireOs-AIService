"""
Excel exporter for document extraction results.

Provides comprehensive Excel export with:
- Multiple worksheet support (data, metadata, validation, audit)
- Conditional formatting and styling
- HIPAA-compliant data handling
- Professional formatting for review workflows
"""

from dataclasses import dataclass, field
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import get_logger
from src.pipeline.state import (
    ExtractionState,
    ExtractionStatus,
)


logger = get_logger(__name__)


class SheetType(str, Enum):
    """Excel sheet types."""

    DATA = "data"
    METADATA = "metadata"
    VALIDATION = "validation"
    AUDIT = "audit"
    PAGE_DETAILS = "page_details"
    RAW_PASSES = "raw_passes"
    PIPELINE = "pipeline"
    # V3 Phase 4 — per-row provenance sheet. Surfaces the FieldValue
    # ``_provenance`` shape (page, bbox, extraction_path,
    # agent_signatures, confidence, vlm_model_id, mem0_match) as a
    # flat sheet keyed by ``field_path``. The data sheet's
    # ``_provenance_ref`` column points back at this sheet's rows so
    # the data sheet stays readable. Only emitted when
    # ``ExcelExportConfig.include_provenance`` is true (auto-true for
    # DETAILED / TECHNICAL styles via :meth:`ExcelExportConfig.from_style`).
    PROVENANCE = "provenance"


class ExcelStyle(str, Enum):
    """Excel export style — mirrors :class:`MarkdownStyle`.

    Determines which sheets ship by default and whether the
    provenance sheet is included. See
    :meth:`ExcelExportConfig.from_style` for the mapping.
    """

    MINIMAL = "minimal"  # Data sheet only
    SUMMARY = "summary"  # Data + metadata
    DETAILED = "detailed"  # Standard 5 sheets + provenance
    TECHNICAL = "technical"  # All sheets + raw passes + provenance


@dataclass(slots=True)
class SheetConfig:
    """
    Configuration for individual Excel sheet.

    Attributes:
        sheet_type: Type of sheet to generate.
        sheet_name: Custom name for the sheet.
        include: Whether to include this sheet.
        freeze_panes: Cell to freeze panes at (e.g., "A2").
        auto_filter: Enable auto-filter on headers.
        column_widths: Custom column widths mapping.
    """

    sheet_type: SheetType
    sheet_name: str | None = None
    include: bool = True
    freeze_panes: str = "A2"
    auto_filter: bool = True
    column_widths: dict[str, int] = field(default_factory=dict)


@dataclass(slots=True)
class ExcelExportConfig:
    """
    Configuration for Excel export.

    Attributes:
        sheets: List of sheet configurations.
        include_styling: Apply professional styling.
        include_confidence_colors: Color-code by confidence level.
        include_validation_highlighting: Highlight validation issues.
        include_formulas: Include Excel formulas for calculations.
        mask_phi: Apply PHI masking to specified fields.
        phi_fields: Fields to mask for PHI compliance.
        phi_mask_pattern: Pattern to use for PHI masking.
        default_column_width: Default column width.
        header_row_height: Height of header rows.
        data_row_height: Height of data rows.
    """

    sheets: list[SheetConfig] = field(
        default_factory=lambda: [
            SheetConfig(SheetType.DATA, "Extracted Data"),
            SheetConfig(SheetType.METADATA, "Processing Metadata"),
            SheetConfig(SheetType.VALIDATION, "Validation Results"),
            SheetConfig(SheetType.AUDIT, "Audit Trail"),
            SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence"),
        ]
    )
    include_styling: bool = True
    include_confidence_colors: bool = True
    include_validation_highlighting: bool = True
    include_formulas: bool = True
    mask_phi: bool = False
    phi_fields: set[str] = field(
        default_factory=lambda: {
            "ssn",
            "social_security",
            "member_id",
            "subscriber_id",
            "patient_account",
            "policy_number",
            "group_number",
        }
    )
    phi_mask_pattern: str = "***MASKED***"
    default_column_width: int = 18
    header_row_height: int = 25
    data_row_height: int = 20
    # V3 Phase 4 — per-row provenance. When ``True`` the exporter:
    # (1) appends a ``Provenance`` worksheet listing every leaf's
    #     ``Provenance`` row (page, bbox, extraction_path, ...), and
    # (2) adds a ``_provenance_ref`` column to the data sheet whose
    #     value is the ``field_path`` key matching the provenance row.
    # Default is ``False`` to preserve the legacy 5-sheet shape that
    # existing callers and tests rely on. :meth:`from_style` flips
    # this on for DETAILED / TECHNICAL.
    include_provenance: bool = False

    @classmethod
    def from_style(
        cls,
        style: "ExcelStyle",
        *,
        include_provenance: bool | None = None,
        **overrides: Any,
    ) -> "ExcelExportConfig":
        """Construct a config from an :class:`ExcelStyle`.

        Mapping (matches the master plan Appendix D contract):

        * ``MINIMAL``   → ``[DATA]`` only, provenance OFF.
        * ``SUMMARY``   → ``[DATA, METADATA]``, provenance OFF.
        * ``DETAILED``  → 5-sheet legacy roster + ``PROVENANCE``;
          provenance ON.
        * ``TECHNICAL`` → all sheets including ``RAW_PASSES`` and
          ``PAGE_DETAILS`` + ``PROVENANCE``; provenance ON.

        ``include_provenance`` overrides the style default. MINIMAL /
        SUMMARY callers must opt in explicitly per the master plan
        ("only include it if explicitly opted in").

        Any keyword in ``overrides`` (e.g. ``mask_phi=True``) is
        forwarded to the constructor unchanged.
        """
        roster_by_style: dict[ExcelStyle, list[SheetConfig]] = {
            ExcelStyle.MINIMAL: [
                SheetConfig(SheetType.DATA, "Extracted Data"),
            ],
            ExcelStyle.SUMMARY: [
                SheetConfig(SheetType.DATA, "Extracted Data"),
                SheetConfig(SheetType.METADATA, "Processing Metadata"),
            ],
            ExcelStyle.DETAILED: [
                SheetConfig(SheetType.DATA, "Extracted Data"),
                SheetConfig(SheetType.METADATA, "Processing Metadata"),
                SheetConfig(SheetType.VALIDATION, "Validation Results"),
                SheetConfig(SheetType.AUDIT, "Audit Trail"),
                SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence"),
            ],
            ExcelStyle.TECHNICAL: [
                SheetConfig(SheetType.DATA, "Extracted Data"),
                SheetConfig(SheetType.METADATA, "Processing Metadata"),
                SheetConfig(SheetType.VALIDATION, "Validation Results"),
                SheetConfig(SheetType.AUDIT, "Audit Trail"),
                SheetConfig(SheetType.PIPELINE, "Pipeline Intelligence"),
                SheetConfig(SheetType.PAGE_DETAILS, "Page Details"),
                SheetConfig(SheetType.RAW_PASSES, "Raw Passes"),
            ],
        }
        sheets = list(roster_by_style[style])
        # Default provenance behaviour per style.
        prov_default = style in (ExcelStyle.DETAILED, ExcelStyle.TECHNICAL)
        prov_on = prov_default if include_provenance is None else include_provenance
        if prov_on:
            sheets.append(SheetConfig(SheetType.PROVENANCE, "Provenance"))
        return cls(sheets=sheets, include_provenance=prov_on, **overrides)


class ExcelStyler:
    """Provides consistent styling for Excel exports."""

    # Color definitions
    HEADER_BG = "1F4E79"
    HEADER_FG = "FFFFFF"
    ALT_ROW_BG = "F2F2F2"
    HIGH_CONFIDENCE_BG = "C6EFCE"
    MEDIUM_CONFIDENCE_BG = "FFEB9C"
    LOW_CONFIDENCE_BG = "FFC7CE"
    ERROR_BG = "FF6B6B"
    WARNING_BG = "FFD93D"
    SUCCESS_BG = "6BCB77"

    def __init__(self) -> None:
        """Initialize the Excel styler with predefined styles."""
        self._header_font = Font(bold=True, color=self.HEADER_FG, size=11)
        self._header_fill = PatternFill(
            start_color=self.HEADER_BG,
            end_color=self.HEADER_BG,
            fill_type="solid",
        )
        self._header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        self._data_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self._alt_row_fill = PatternFill(
            start_color=self.ALT_ROW_BG,
            end_color=self.ALT_ROW_BG,
            fill_type="solid",
        )

    def apply_header_style(self, cell: Any) -> None:
        """Apply header styling to a cell."""
        cell.font = self._header_font
        cell.fill = self._header_fill
        cell.alignment = self._header_alignment
        cell.border = self._thin_border

    def apply_data_style(self, cell: Any, row_index: int) -> None:
        """Apply data cell styling with alternating row colors."""
        cell.alignment = self._data_alignment
        cell.border = self._thin_border
        if row_index % 2 == 0:
            cell.fill = self._alt_row_fill

    def apply_confidence_color(self, cell: Any, confidence: float) -> None:
        """Apply color based on confidence level."""
        # Thresholds: HIGH >= 0.85, MEDIUM >= 0.50, LOW < 0.50
        if confidence >= 0.85:
            fill_color = self.HIGH_CONFIDENCE_BG
        elif confidence >= 0.50:
            fill_color = self.MEDIUM_CONFIDENCE_BG
        else:
            fill_color = self.LOW_CONFIDENCE_BG

        cell.fill = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type="solid",
        )

    def apply_status_color(self, cell: Any, status: str) -> None:
        """Apply color based on status value."""
        status_colors = {
            ExtractionStatus.COMPLETED.value: self.SUCCESS_BG,
            ExtractionStatus.FAILED.value: self.ERROR_BG,
            ExtractionStatus.HUMAN_REVIEW.value: self.WARNING_BG,
        }
        color = status_colors.get(status)
        if color:
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid",
            )

    def apply_validation_color(self, cell: Any, is_valid: bool) -> None:
        """Apply color based on validation status."""
        color = self.SUCCESS_BG if is_valid else self.ERROR_BG
        cell.fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid",
        )


class ExcelExporter:
    """
    Export extraction results to Excel format.

    Supports multiple worksheets with comprehensive formatting
    for professional review workflows.
    """

    def __init__(self, config: ExcelExportConfig | None = None) -> None:
        """
        Initialize the Excel exporter.

        Args:
            config: Export configuration (uses defaults if not provided).
        """
        self.config = config or ExcelExportConfig()
        self._styler = ExcelStyler()
        self._logger = logger

    def export(
        self,
        state: ExtractionState,
        output_path: Path | str,
    ) -> Path:
        """
        Export extraction state to Excel file.

        Args:
            state: Extraction state to export.
            output_path: File path to write output.

        Returns:
            Path to the created Excel file.
        """
        output_path = Path(output_path)
        self._logger.debug(
            "excel_export_start",
            processing_id=state.get("processing_id", ""),
            output_path=str(output_path),
        )

        workbook = Workbook()
        # Remove default sheet
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)

        # Build sheets based on configuration
        for sheet_config in self.config.sheets:
            if not sheet_config.include:
                continue

            sheet_name = sheet_config.sheet_name or sheet_config.sheet_type.value
            worksheet = workbook.create_sheet(title=sheet_name)

            self._build_sheet(worksheet, sheet_config, state)

        # Write to file
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        self._logger.info(
            "excel_file_written",
            path=str(output_path),
            sheet_count=len(workbook.sheetnames),
        )

        return output_path

    def _build_sheet(
        self,
        worksheet: Worksheet,
        config: SheetConfig,
        state: ExtractionState,
    ) -> None:
        """Build a worksheet based on its type."""
        builders = {
            SheetType.DATA: self._build_data_sheet,
            SheetType.METADATA: self._build_metadata_sheet,
            SheetType.VALIDATION: self._build_validation_sheet,
            SheetType.AUDIT: self._build_audit_sheet,
            SheetType.PAGE_DETAILS: self._build_page_details_sheet,
            SheetType.RAW_PASSES: self._build_raw_passes_sheet,
            SheetType.PIPELINE: self._build_pipeline_sheet,
            SheetType.PROVENANCE: self._build_provenance_sheet,
        }

        builder = builders.get(config.sheet_type)
        if builder:
            builder(worksheet, state)

        # Apply sheet-level settings
        if config.freeze_panes:
            worksheet.freeze_panes = config.freeze_panes

        if config.auto_filter and worksheet.max_row and worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

    def _build_data_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the main data extraction sheet."""
        headers = [
            "Field Name",
            "Value",
            "Confidence",
            "Confidence Level",
            "Location",
            "Passes Agree",
            "Bbox Page",
            "Bbox X",
            "Bbox Y",
            "Bbox W",
            "Bbox H",
        ]
        # V3 Phase 4 — when provenance is enabled the data sheet
        # carries a back-reference into the Provenance sheet (matched
        # by ``field_path``). We keep the bbox columns above for
        # legacy callers who scrape coordinates from the data sheet
        # directly; the new column is purely additive.
        if self.config.include_provenance:
            headers.append("_provenance_ref")
        self._write_header_row(worksheet, headers)

        # Calculate column index dynamically to avoid hardcoding
        # Column indices are 1-based in openpyxl
        confidence_column_idx = headers.index("Confidence") + 1

        merged = state.get("merged_extraction", {})
        field_meta = state.get("field_metadata", {})
        # Field paths covered by the provenance sheet (merged_extraction_v2
        # is the FieldValue-shaped twin; the provenance sheet keys on
        # those names). Empty when v2 isn't populated.
        merged_v2 = state.get("merged_extraction_v2") or {}
        provenance_field_paths: set[str] = (
            set(merged_v2.keys()) if isinstance(merged_v2, dict) else set()
        )

        row = 2
        for field_name, field_data in merged.items():
            if field_name.lower() in self.config.phi_fields and self.config.mask_phi:
                value = self._mask_phi_value(
                    field_data.get("value") if isinstance(field_data, dict) else field_data
                )
            elif isinstance(field_data, dict):
                value = field_data.get("value", "")
            else:
                value = field_data

            meta = field_meta.get(field_name, {})
            confidence = meta.get("confidence", 0.0) if isinstance(meta, dict) else 0.0
            confidence_level = (
                meta.get("confidence_level", "low") if isinstance(meta, dict) else "low"
            )
            location = field_data.get("location", "") if isinstance(field_data, dict) else ""
            passes_agree = meta.get("passes_agree", True) if isinstance(meta, dict) else True

            # Extract bbox data if present
            bbox = meta.get("bbox", {}) if isinstance(meta, dict) else {}
            bbox_page = bbox.get("page", "") if bbox else ""
            bbox_x = bbox.get("x", "") if bbox else ""
            bbox_y = bbox.get("y", "") if bbox else ""
            bbox_w = bbox.get("width", "") if bbox else ""
            bbox_h = bbox.get("height", "") if bbox else ""

            row_data = [
                field_name,
                str(value) if value is not None else "",
                confidence,
                confidence_level,
                location,
                "Yes" if passes_agree else "No",
                bbox_page,
                bbox_x,
                bbox_y,
                bbox_w,
                bbox_h,
            ]
            if self.config.include_provenance:
                # Reference points to the Provenance sheet's
                # ``field_path`` cell. Only emit when v2 actually
                # carries provenance for the field; otherwise leave
                # blank so consumers don't chase a dead link.
                row_data.append(
                    field_name if field_name in provenance_field_paths else ""
                )

            self._write_data_row(worksheet, row, row_data)

            # Apply confidence coloring using dynamically calculated column index
            if self.config.include_confidence_colors:
                self._styler.apply_confidence_color(
                    worksheet.cell(row=row, column=confidence_column_idx), confidence
                )

            row += 1

        self._auto_size_columns(worksheet)

    def _build_metadata_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the processing metadata sheet."""
        headers = ["Property", "Value"]
        self._write_header_row(worksheet, headers)

        metadata = [
            ("Processing ID", state.get("processing_id", "")),
            ("PDF Path", state.get("pdf_path", "")),
            ("PDF Hash", state.get("pdf_hash", "")),
            ("Document Type", state.get("document_type", "")),
            ("Schema Name", state.get("selected_schema_name", "")),
            ("Status", state.get("status", "")),
            ("Page Count", len(state.get("page_images", []))),
            ("Start Time", state.get("start_time", "")),
            ("End Time", state.get("end_time", "")),
            ("Total VLM Calls", state.get("total_vlm_calls", 0)),
            ("Processing Time (ms)", state.get("total_processing_time_ms", 0)),
            ("Retry Count", state.get("retry_count", 0)),
            ("Overall Confidence", state.get("overall_confidence", 0.0)),
            ("Confidence Level", state.get("confidence_level", "")),
            ("Requires Human Review", state.get("requires_human_review", False)),
            ("Human Review Reason", state.get("human_review_reason", "")),
        ]

        for row, (prop, value) in enumerate(metadata, start=2):
            row_data = [prop, str(value) if value is not None else ""]
            self._write_data_row(worksheet, row, row_data)

            # Apply status coloring
            if prop == "Status" and self.config.include_styling:
                self._styler.apply_status_color(worksheet.cell(row=row, column=2), str(value))

        self._auto_size_columns(worksheet)

    def _build_validation_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the validation results sheet."""
        headers = ["Field Name", "Is Valid", "Validation Type", "Message", "Severity"]
        self._write_header_row(worksheet, headers)

        validation = state.get("validation", {})
        field_validations = validation.get("field_validations", {})

        row = 2
        for field_name, field_val in field_validations.items():
            if isinstance(field_val, dict):
                is_valid = field_val.get("is_valid", True)
                val_type = field_val.get("validation_type", "")
                message = field_val.get("message", "")
                severity = field_val.get("severity", "info")
            else:
                is_valid = True
                val_type = ""
                message = ""
                severity = "info"

            row_data = [field_name, "Pass" if is_valid else "Fail", val_type, message, severity]
            self._write_data_row(worksheet, row, row_data)

            if self.config.include_validation_highlighting:
                self._styler.apply_validation_color(worksheet.cell(row=row, column=2), is_valid)

            row += 1

        # Add cross-field validations
        cross_field = validation.get("cross_field_validations", [])
        for cf_val in cross_field:
            if isinstance(cf_val, dict):
                row_data = [
                    "Cross-Field",
                    "Pass" if cf_val.get("is_valid", True) else "Fail",
                    cf_val.get("rule_name", ""),
                    cf_val.get("message", ""),
                    cf_val.get("severity", "info"),
                ]
                self._write_data_row(worksheet, row, row_data)
                row += 1

        # Add hallucination flags
        hallucination_flags = validation.get("hallucination_flags", [])
        for flag in hallucination_flags:
            if isinstance(flag, dict):
                row_data = [
                    flag.get("field_name", "Unknown"),
                    "Fail",
                    "Hallucination Detection",
                    flag.get("reason", ""),
                    "critical",
                ]
                self._write_data_row(worksheet, row, row_data)
                if self.config.include_validation_highlighting:
                    self._styler.apply_validation_color(worksheet.cell(row=row, column=2), False)
                row += 1

        self._auto_size_columns(worksheet)

    def _build_audit_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the audit trail sheet."""
        headers = ["Timestamp", "Event", "Details"]
        self._write_header_row(worksheet, headers)

        audit_events = [
            (
                state.get("start_time", ""),
                "Processing Started",
                f"PDF: {state.get('pdf_path', '')}",
            ),
            (state.get("start_time", ""), "Document Type Detected", state.get("document_type", "")),
            (state.get("start_time", ""), "Schema Selected", state.get("selected_schema_name", "")),
        ]

        # Add page extraction events
        for page in state.get("page_extractions", []):
            if isinstance(page, dict):
                page_num = page.get("page_number", 0)
                extraction_time = page.get("extraction_time_ms", 0)
                audit_events.append(
                    (
                        "",
                        f"Page {page_num} Extracted",
                        f"Fields: {len(page.get('merged_fields', {}))}, Time: {extraction_time}ms",
                    )
                )

        # Add validation event
        validation = state.get("validation", {})
        if validation:
            is_valid = validation.get("is_valid", False)
            audit_events.append(
                (
                    "",
                    "Validation Completed",
                    f"Result: {'Pass' if is_valid else 'Fail'}",
                )
            )

        # Add completion/failure event
        status = state.get("status", "")
        end_time = state.get("end_time", "")
        if status == ExtractionStatus.COMPLETED.value:
            audit_events.append((end_time, "Processing Completed", "Success"))
        elif status == ExtractionStatus.FAILED.value:
            errors = state.get("errors", [])
            error_msg = errors[0] if errors else "Unknown error"
            audit_events.append((end_time, "Processing Failed", str(error_msg)))
        elif status == ExtractionStatus.HUMAN_REVIEW.value:
            reason = state.get("human_review_reason", "")
            audit_events.append((end_time, "Requires Human Review", reason))

        # Add export timestamp
        audit_events.append(
            (
                datetime.now(UTC).isoformat(),
                "Excel Export Generated",
                "Export timestamp",
            )
        )

        for row, (timestamp, event, details) in enumerate(audit_events, start=2):
            row_data = [str(timestamp), event, details]
            self._write_data_row(worksheet, row, row_data)

        self._auto_size_columns(worksheet)

    def _build_page_details_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the page-level details sheet."""
        headers = [
            "Page Number",
            "Field Count",
            "Confidence",
            "Agreement Rate",
            "VLM Calls",
            "Extraction Time (ms)",
            "Errors",
        ]
        self._write_header_row(worksheet, headers)

        pages = state.get("page_extractions", [])
        for row, page in enumerate(pages, start=2):
            if isinstance(page, dict):
                errors = page.get("errors", [])
                error_str = "; ".join(str(e) for e in errors) if errors else ""

                row_data = [
                    page.get("page_number", 0),
                    len(page.get("merged_fields", {})),
                    page.get("overall_confidence", 0.0),
                    page.get("agreement_rate", 0.0),
                    page.get("vlm_calls", 0),
                    page.get("extraction_time_ms", 0),
                    error_str,
                ]
                self._write_data_row(worksheet, row, row_data)

                if self.config.include_confidence_colors:
                    confidence = page.get("overall_confidence", 0.0)
                    self._styler.apply_confidence_color(
                        worksheet.cell(row=row, column=3), confidence
                    )

        self._auto_size_columns(worksheet)

    def _build_raw_passes_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the raw pass data sheet."""
        headers = ["Page", "Pass", "Field", "Value", "Confidence"]
        self._write_header_row(worksheet, headers)

        pages = state.get("page_extractions", [])
        row = 2

        for page in pages:
            if not isinstance(page, dict):
                continue

            page_num = page.get("page_number", 0)

            # Pass 1 data
            pass1 = page.get("pass1_raw", {})
            for field_name, field_data in pass1.items():
                if isinstance(field_data, dict):
                    value = field_data.get("value", "")
                    confidence = field_data.get("confidence", 0.0)
                else:
                    value = field_data
                    confidence = 0.0

                row_data = [page_num, "Pass 1", field_name, str(value), confidence]
                self._write_data_row(worksheet, row, row_data)
                row += 1

            # Pass 2 data
            pass2 = page.get("pass2_raw", {})
            for field_name, field_data in pass2.items():
                if isinstance(field_data, dict):
                    value = field_data.get("value", "")
                    confidence = field_data.get("confidence", 0.0)
                else:
                    value = field_data
                    confidence = 0.0

                row_data = [page_num, "Pass 2", field_name, str(value), confidence]
                self._write_data_row(worksheet, row, row_data)
                row += 1

        self._auto_size_columns(worksheet)

    def _build_pipeline_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """Build the pipeline intelligence sheet with Phase 2A-3C metadata."""
        headers = ["Category", "Property", "Value"]
        self._write_header_row(worksheet, headers)

        rows: list[tuple[str, str, Any]] = []

        # --- Document Splitting (Phase 2A) ---
        is_multi = state.get("is_multi_document", False)
        segments = state.get("document_segments", [])
        rows.append(("Document Splitting", "Is Multi-Document", "Yes" if is_multi else "No"))
        rows.append(("Document Splitting", "Segment Count", len(segments)))
        for i, seg in enumerate(segments):
            if isinstance(seg, dict):
                start = seg.get("start_page", "?")
                end = seg.get("end_page", "?")
                doc_type = seg.get("document_type", "unknown")
                rows.append(
                    ("Document Splitting", f"Segment {i + 1}", f"Pages {start}-{end} ({doc_type})")
                )

        # --- Table Detection (Phase 2B) ---
        tables = state.get("detected_tables", [])
        rows.append(("Table Detection", "Tables Detected", len(tables)))
        for i, tbl in enumerate(tables):
            if isinstance(tbl, dict):
                page = tbl.get("page", "?")
                row_count = tbl.get("row_count", tbl.get("rows", "?"))
                col_count = tbl.get("column_count", tbl.get("columns", "?"))
                rows.append(
                    ("Table Detection", f"Table {i + 1}", f"Page {page}: {row_count}R x {col_count}C")
                )

        # --- Schema Proposal (Phase 2C) ---
        proposal = state.get("schema_proposal")
        rows.append(("Schema Proposal", "Proposal Generated", "Yes" if proposal else "No"))
        if isinstance(proposal, dict):
            rows.append(
                ("Schema Proposal", "Proposed Schema Name", proposal.get("schema_name", "N/A"))
            )
            rows.append(
                ("Schema Proposal", "Proposed Field Count", len(proposal.get("fields", [])))
            )

        # --- Dynamic Prompt Enhancement (Phase 3B) ---
        enhancement = state.get("prompt_enhancement_applied", False)
        rows.append(
            ("Prompt Enhancement", "Correction-Based Enhancement", "Applied" if enhancement else "Not Applied")
        )

        # --- Adaptive Extraction ---
        adaptive = state.get("use_adaptive_extraction", False)
        rows.append(("Extraction Mode", "Adaptive (VLM-First)", "Yes" if adaptive else "No (Legacy)"))

        layout_count = len(state.get("layout_analyses", []))
        component_count = len(state.get("component_maps", []))
        rows.append(("Extraction Mode", "Layout Analyses", layout_count))
        rows.append(("Extraction Mode", "Component Maps", component_count))

        has_adaptive_schema = state.get("adaptive_schema") is not None
        rows.append(
            ("Extraction Mode", "Adaptive Schema Generated", "Yes" if has_adaptive_schema else "No")
        )

        # --- Memory Context ---
        similar = state.get("similar_docs", [])
        rows.append(("Memory", "Similar Documents Found", len(similar)))
        rows.append(
            ("Memory", "Correction Hints Available", "Yes" if state.get("correction_hints") else "No")
        )
        rows.append(
            ("Memory", "Provider Patterns Available", "Yes" if state.get("provider_patterns") else "No")
        )

        # Write all rows
        for row_idx, (category, prop, value) in enumerate(rows, start=2):
            row_data = [category, prop, str(value) if value is not None else ""]
            self._write_data_row(worksheet, row_idx, row_data)

        self._auto_size_columns(worksheet)

    def _build_provenance_sheet(
        self,
        worksheet: Worksheet,
        state: ExtractionState,
    ) -> None:
        """V3 Phase 4 — build the per-row provenance worksheet.

        Mirrors :meth:`JSONExporter._build_provenance_block`: walks
        ``merged_extraction_v2`` (the FieldValue-shaped twin), pulls
        the ``Provenance`` off each leaf via :func:`unwrap_provenance`,
        and writes one row per field. The data sheet's
        ``_provenance_ref`` column carries the matching ``field_path``
        key so reviewers can cross-reference without baking bbox
        coordinates into every data row.

        Empty (header only) when ``merged_extraction_v2`` isn't
        populated — same back-compat behaviour as the JSON exporter's
        provenance block.
        """
        from src.pipeline.provenance import unwrap_provenance

        headers = [
            "field_path",
            "page",
            "bbox_x",
            "bbox_y",
            "bbox_width",
            "bbox_height",
            "source_block_id",
            "extraction_path",
            "agent_signatures",
            "confidence",
            "vlm_model_id",
            "mem0_match",
        ]
        self._write_header_row(worksheet, headers)

        merged_v2 = state.get("merged_extraction_v2") or {}
        if not isinstance(merged_v2, dict) or not merged_v2:
            self._auto_size_columns(worksheet)
            return

        confidence_column_idx = headers.index("confidence") + 1

        row = 2
        for field_name, wrapper in merged_v2.items():
            if field_name in self.config.phi_fields and self.config.mask_phi:
                # PHI redaction policy: keep the structured provenance
                # (page/bbox are about the extraction act, not the
                # patient) but redact the field_path so the sheet can
                # be shared with non-privileged reviewers without
                # leaking the field name's PHI association. See
                # PHIFieldValue docstring for the broader principle.
                field_path_cell: Any = self.config.phi_mask_pattern
            else:
                field_path_cell = field_name

            prov = unwrap_provenance(wrapper)
            if prov is None:
                # No provenance available — emit a placeholder row so
                # the data-sheet's _provenance_ref doesn't dangle.
                row_data = [field_path_cell, "", "", "", "", "", "", "", "", "", "", ""]
                self._write_data_row(worksheet, row, row_data)
                row += 1
                continue

            serial = prov.to_serialisable()
            bbox = serial.get("bbox") or {}
            bbox_x = bbox.get("x", "") if isinstance(bbox, dict) else ""
            bbox_y = bbox.get("y", "") if isinstance(bbox, dict) else ""
            bbox_w = bbox.get("width", "") if isinstance(bbox, dict) else ""
            bbox_h = bbox.get("height", "") if isinstance(bbox, dict) else ""

            # Lists get joined with "; " so the cell renders cleanly
            # in Excel while preserving order. Reviewers who need the
            # structured form can read the JSON export instead.
            ext_path = serial.get("extraction_path") or []
            agent_sigs = serial.get("agent_signatures") or []

            confidence = float(serial.get("confidence") or 0.0)

            row_data = [
                field_path_cell,
                serial.get("page", ""),
                bbox_x,
                bbox_y,
                bbox_w,
                bbox_h,
                serial.get("source_block_id", ""),
                "; ".join(str(s) for s in ext_path),
                "; ".join(str(s) for s in agent_sigs),
                confidence,
                serial.get("vlm_model_id", ""),
                serial.get("mem0_match") or "",
            ]
            self._write_data_row(worksheet, row, row_data)

            if self.config.include_confidence_colors:
                self._styler.apply_confidence_color(
                    worksheet.cell(row=row, column=confidence_column_idx),
                    confidence,
                )

            row += 1

        self._auto_size_columns(worksheet)

    def _write_header_row(self, worksheet: Worksheet, headers: list[str]) -> None:
        """Write and style header row."""
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            if self.config.include_styling:
                self._styler.apply_header_style(cell)

        worksheet.row_dimensions[1].height = self.config.header_row_height

    def _write_data_row(
        self,
        worksheet: Worksheet,
        row: int,
        values: list[Any],
    ) -> None:
        """Write and style a data row."""
        for col, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row, column=col, value=value)
            if self.config.include_styling:
                self._styler.apply_data_style(cell, row)

        worksheet.row_dimensions[row].height = self.config.data_row_height

    def _auto_size_columns(self, worksheet: Worksheet) -> None:
        """Auto-size columns based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    # Use 'is not None' to preserve numeric zero values (e.g., "0" for counts)
                    cell_value = str(cell.value) if cell.value is not None else ""
                    cell_length = len(cell_value)
                    max_length = max(max_length, cell_length)
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, self.config.default_column_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _mask_phi_value(self, value: Any) -> str:
        """Mask a PHI value."""
        if value is None:
            return self.config.phi_mask_pattern
        if isinstance(value, str) and len(value) > 4:
            return f"{value[:2]}{self.config.phi_mask_pattern}{value[-2:]}"
        return self.config.phi_mask_pattern


def export_to_excel(
    state: ExtractionState,
    output_path: Path | str,
    include_styling: bool = True,
    include_confidence_colors: bool = True,
    include_validation_highlighting: bool = True,
    mask_phi: bool = False,
) -> Path:
    """
    Convenience function to export extraction state to Excel.

    Args:
        state: Extraction state to export.
        output_path: File path to write output.
        include_styling: Apply professional styling.
        include_confidence_colors: Color-code by confidence level.
        include_validation_highlighting: Highlight validation issues.
        mask_phi: Apply PHI masking to sensitive fields.

    Returns:
        Path to the created Excel file.

    Example:
        >>> path = export_to_excel(state, "output.xlsx", mask_phi=True)
        >>> print(f"Excel file created at: {path}")
    """
    config = ExcelExportConfig(
        include_styling=include_styling,
        include_confidence_colors=include_confidence_colors,
        include_validation_highlighting=include_validation_highlighting,
        mask_phi=mask_phi,
    )

    exporter = ExcelExporter(config)
    return exporter.export(state, output_path)
