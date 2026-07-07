"""
Consolidated export for multi-record extraction results.

Provides:
- Cross-page duplicate detection
- Excel export with All Records, Duplicates, Page Summary, Processing Summary sheets
- Markdown export with per-record sections
- JSON export
- Record completeness validation
"""

import copy
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import get_logger
from src.extraction.multi_record import DocumentExtractionResult, ExtractedRecord
from src.security.phi_mask import enforce_mask_phi


logger = get_logger(__name__)


def _apply_mask_phi(
    result: DocumentExtractionResult,
    *,
    mask_phi: bool,
) -> DocumentExtractionResult:
    """Return a deep-copied result with PHI fields redacted if requested.

    Single chokepoint for export-time PHI masking. Defers to
    ``src.security.phi_mask.enforce_mask_phi`` so all formats stay
    consistent. The input is never mutated.
    """
    if not mask_phi:
        return result

    masked = copy.deepcopy(result)
    for rec in masked.records:
        # ExtractedRecord.fields is a dict[str, Any].
        rec.fields = enforce_mask_phi(rec.fields)
        # primary_identifier is often a patient-name-like string.
        rec.primary_identifier = "[REDACTED]"
    return masked


def detect_duplicates(
    records: list[ExtractedRecord],
) -> dict[str, list[int]]:
    """
    Detect duplicate records across pages using primary identifier similarity.

    Returns:
        Dict mapping normalized identifier to list of record indices.
    """
    identifier_groups: dict[str, list[int]] = defaultdict(list)

    for idx, record in enumerate(records):
        normalized = " ".join(record.primary_identifier.lower().strip().split())
        identifier_groups[normalized].append(idx)

    duplicates = {
        ident: indices
        for ident, indices in identifier_groups.items()
        if len(indices) > 1
    }

    if duplicates:
        logger.info(
            "duplicates_detected",
            count=len(duplicates),
            identifiers=list(duplicates.keys()),
        )
    else:
        logger.info("no_duplicates_found")

    return duplicates


def validate_record_completeness(
    record: ExtractedRecord, schema: dict[str, Any]
) -> dict[str, Any]:
    """Validate record completeness against schema."""
    fields = record.fields
    schema_fields = schema.get("fields", [])
    expected_count = len(schema_fields)

    required_fields = [f["field_name"] for f in schema_fields if f.get("required")]
    missing_required = []
    for fname in required_fields:
        val = fields.get(fname)
        if val is None or val == "" or val == []:
            missing_required.append(fname)

    empty_fields = []
    for fname, val in fields.items():
        if val is None or val == "" or val == []:
            empty_fields.append(fname)

    non_empty = len(fields) - len(empty_fields)
    completeness = non_empty / max(expected_count, 1)

    return {
        "is_complete": len(missing_required) == 0,
        "missing_required_fields": missing_required,
        "empty_fields": empty_fields,
        "field_count": len(fields),
        "expected_field_count": expected_count,
        "completeness_score": completeness,
    }


def export_excel(
    result: DocumentExtractionResult,
    output_path: str | Path,
    *,
    mask_phi: bool = False,
) -> None:
    """
    Export extraction results to a consolidated Excel workbook.

    Sheets:
    - All Records: One row per record with all fields
    - Duplicates: Cross-page duplicate analysis
    - Page Summary: Per-page statistics
    - Processing Summary: Overall extraction metrics

    Args:
        mask_phi: If True, route records through
            ``src.security.phi_mask.enforce_mask_phi`` before rendering.
            PHI field names and PHI-shaped values become ``[REDACTED]``.
    """
    output_path = Path(output_path)
    logger.info("exporting_excel", output_path=str(output_path), mask_phi=mask_phi)
    result = _apply_mask_phi(result, mask_phi=mask_phi)

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    dup_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")

    records = result.records
    schema = result.schema

    # Detect duplicates
    duplicates = detect_duplicates(records)
    duplicate_indices: set[int] = set()
    for indices in duplicates.values():
        duplicate_indices.update(indices)

    # Validate records
    validations = [validate_record_completeness(r, schema) for r in records]

    # ── Sheet 1: All Records ──
    ws = wb.create_sheet("All Records")
    headers = ["Record #", "Page", "Primary ID"]
    for f in schema.get("fields", []):
        headers.append(f.get("display_name", f["field_name"]))
    headers.extend(["Confidence", "Is Duplicate", "Completeness"])

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    for ri, (rec, val) in enumerate(zip(records, validations, strict=False), 2):
        is_dup = (ri - 2) in duplicate_indices
        row = [rec.record_id, rec.page_number, rec.primary_identifier]

        for f in schema.get("fields", []):
            v = rec.fields.get(f["field_name"], "")
            if isinstance(v, list):
                v = "\n".join(str(x) for x in v)
            row.append(str(v) if v else "")

        row.extend([
            f"{rec.confidence:.0%}",
            "YES" if is_dup else "NO",
            f"{val['completeness_score']:.0%}",
        ])

        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border
            if is_dup:
                cell.fill = dup_fill
            elif ri % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[ri].height = 30

    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Duplicates ──
    if duplicates:
        ws_d = wb.create_sheet("Duplicates")
        d_headers = ["Primary Identifier", "Occurrences", "Pages", "Record IDs", "Action"]
        for ci, h in enumerate(d_headers, 1):
            cell = ws_d.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ri = 2
        for ident, indices in duplicates.items():
            pages = [records[i].page_number for i in indices]
            rec_ids = [records[i].record_id for i in indices]
            row = [
                ident.title(),
                len(indices),
                ", ".join(map(str, pages)),
                ", ".join(map(str, rec_ids)),
                "Review and merge if same entity",
            ]
            for ci, v in enumerate(row, 1):
                cell = ws_d.cell(row=ri, column=ci, value=v)
                cell.fill = warn_fill
            ri += 1

        for ci, w in enumerate([30, 12, 15, 15, 40], 1):
            ws_d.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Page Summary ──
    ws_p = wb.create_sheet("Page Summary")
    p_headers = ["Page", "Records", "Avg Confidence", "Unique IDs", "Duplicates"]
    for ci, h in enumerate(p_headers, 1):
        cell = ws_p.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    page_stats: dict[int, dict[str, Any]] = defaultdict(
        lambda: {"records": [], "identifiers": set()}
    )
    for rec in records:
        page_stats[rec.page_number]["records"].append(rec)
        page_stats[rec.page_number]["identifiers"].add(
            rec.primary_identifier.lower()
        )

    ri = 2
    for pn in sorted(page_stats.keys()):
        st = page_stats[pn]
        avg_c = sum(r.confidence for r in st["records"]) / len(st["records"])
        page_rec_indices = [
            i for i, r in enumerate(records) if r.page_number == pn
        ]
        page_dups = sum(1 for i in page_rec_indices if i in duplicate_indices)

        row = [pn, len(st["records"]), f"{avg_c:.0%}", len(st["identifiers"]), page_dups]
        for ci, v in enumerate(row, 1):
            ws_p.cell(row=ri, column=ci, value=v)
        ri += 1

    for ci in range(1, len(p_headers) + 1):
        ws_p.column_dimensions[get_column_letter(ci)].width = 18

    # ── Sheet 4: Processing Summary ──
    ws_s = wb.create_sheet("Processing Summary")
    avg_conf = sum(r.confidence for r in records) / max(len(records), 1)
    summary_rows = [
        ("Document Type", result.document_type),
        ("Entity Type", result.entity_type),
        ("PDF Path", result.pdf_path),
        ("Total Pages", result.total_pages),
        ("Total Records", result.total_records),
        ("Unique Records", result.total_records - len(duplicate_indices)),
        ("Duplicate Records", len(duplicate_indices)),
        (
            "Unique Identifiers",
            len({r.primary_identifier.lower() for r in records}),
        ),
        ("Avg Records/Page", f"{result.total_records / max(result.total_pages, 1):.1f}"),
        ("Avg Confidence", f"{avg_conf:.0%}"),
        ("Processing Time (s)", f"{result.total_processing_time_ms / 1000:.1f}"),
        (
            "Avg Time/Record (s)",
            f"{result.total_processing_time_ms / 1000 / max(result.total_records, 1):.1f}",
        ),
        ("Total VLM Calls", result.total_vlm_calls),
        ("Schema Fields", len(schema.get("fields", []))),
    ]

    ws_s.cell(row=1, column=1, value="Metric").font = header_font
    ws_s.cell(row=1, column=2, value="Value").font = header_font
    ws_s.cell(row=1, column=1).fill = header_fill
    ws_s.cell(row=1, column=2).fill = header_fill

    for ri, (metric, value) in enumerate(summary_rows, 2):
        ws_s.cell(row=ri, column=1, value=metric).font = Font(bold=True)
        ws_s.cell(row=ri, column=2, value=value)

    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 40

    # ── Sheet 5 (V3 Phase 4): Provenance ──
    # Emitted only when at least one record carries ``field_provenance``.
    # Keyed by ``(record_id, field_name)`` so audit consumers can
    # join against the All Records sheet by record_id.
    has_provenance = any(
        getattr(r, "field_provenance", None) for r in records
    )
    if has_provenance:
        ws_pv = wb.create_sheet("Provenance")
        prov_headers = [
            "Record #",
            "Field",
            "Page",
            "Bbox X",
            "Bbox Y",
            "Bbox W",
            "Bbox H",
            "Source Block",
            "Extraction Path",
            "Agent Signatures",
            "Confidence",
            "VLM Model",
            "Mem0 Match",
        ]
        for ci, h in enumerate(prov_headers, 1):
            cell = ws_pv.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws_pv.row_dimensions[1].height = 25
        ws_pv.freeze_panes = "A2"

        ri = 2
        for rec in records:
            field_prov = getattr(rec, "field_provenance", None) or {}
            if not isinstance(field_prov, dict) or not field_prov:
                continue
            for field_name, prov in field_prov.items():
                if not isinstance(prov, dict):
                    continue
                bbox = prov.get("bbox") or {}
                ws_pv.cell(row=ri, column=1, value=rec.record_id)
                ws_pv.cell(row=ri, column=2, value=field_name)
                ws_pv.cell(row=ri, column=3, value=prov.get("page"))
                if isinstance(bbox, dict):
                    ws_pv.cell(row=ri, column=4, value=bbox.get("x"))
                    ws_pv.cell(row=ri, column=5, value=bbox.get("y"))
                    ws_pv.cell(
                        row=ri, column=6, value=bbox.get("width", bbox.get("w"))
                    )
                    ws_pv.cell(
                        row=ri, column=7, value=bbox.get("height", bbox.get("h"))
                    )
                ws_pv.cell(row=ri, column=8, value=prov.get("source_block_id", ""))
                ws_pv.cell(
                    row=ri,
                    column=9,
                    value=",".join(prov.get("extraction_path") or []),
                )
                ws_pv.cell(
                    row=ri,
                    column=10,
                    value=",".join(prov.get("agent_signatures") or []),
                )
                ws_pv.cell(row=ri, column=11, value=prov.get("confidence", 0.0))
                ws_pv.cell(row=ri, column=12, value=prov.get("vlm_model_id", ""))
                ws_pv.cell(row=ri, column=13, value=prov.get("mem0_match"))
                ri += 1

        for col_letter in "ABCDEFGHIJKLM":
            ws_pv.column_dimensions[col_letter].width = 18

    # Save
    wb.save(str(output_path))
    logger.info("excel_exported", path=str(output_path), sheets=len(wb.sheetnames))


def export_markdown(
    result: DocumentExtractionResult,
    output_path: str | Path,
    *,
    mask_phi: bool = False,
) -> None:
    """
    Export extraction results to a consolidated Markdown report.

    Sections:
    - Document summary
    - Duplicate warnings
    - Page-level summary table
    - Per-record detail sections

    Args:
        mask_phi: If True, redact PHI fields/values before rendering.
    """
    output_path = Path(output_path)
    logger.info("exporting_markdown", output_path=str(output_path), mask_phi=mask_phi)
    result = _apply_mask_phi(result, mask_phi=mask_phi)

    records = result.records
    schema = result.schema
    duplicates = detect_duplicates(records)

    lines: list[str] = []

    # Title
    title = result.document_type.replace("_", " ").title()
    lines.append(f"# {title} Extraction Report")
    lines.append("")
    lines.append(f"**Document**: {result.pdf_path}")
    lines.append(f"**Pages**: {result.total_pages}")
    lines.append(f"**Total Records**: {result.total_records}")
    unique_count = result.total_records - sum(
        len(v) for v in duplicates.values()
    ) + len(duplicates)
    lines.append(f"**Unique Records**: {unique_count}")
    lines.append(f"**Entity Type**: {result.entity_type}")
    lines.append(
        f"**Processing Time**: {result.total_processing_time_ms / 1000:.1f}s"
    )
    lines.append(f"**VLM Calls**: {result.total_vlm_calls}")
    lines.append("")

    # Duplicates
    if duplicates:
        lines.append("## Duplicate Records Detected")
        lines.append("")
        lines.append(f"Found {len(duplicates)} entities appearing multiple times:")
        lines.append("")
        for ident, indices in duplicates.items():
            pages = [records[i].page_number for i in indices]
            lines.append(
                f"- **{ident.title()}**: Pages {', '.join(map(str, pages))}"
            )
        lines.append("")

    # Page summary table
    lines.append("## Summary by Page")
    lines.append("")
    lines.append("| Page | Records | Avg Confidence |")
    lines.append("|------|---------|----------------|")

    page_groups: dict[int, list[ExtractedRecord]] = defaultdict(list)
    for rec in records:
        page_groups[rec.page_number].append(rec)

    for pn in sorted(page_groups.keys()):
        pr = page_groups[pn]
        avg_c = sum(r.confidence for r in pr) / len(pr)
        lines.append(f"| {pn} | {len(pr)} | {avg_c:.0%} |")
    lines.append("")

    # All records
    lines.append("## All Records")
    lines.append("")

    for rec in records:
        lines.append(
            f"### Record {rec.record_id} - Page {rec.page_number}"
        )
        lines.append(
            f"**{result.entity_type.title()}**: {rec.primary_identifier}"
        )
        lines.append(f"**Confidence**: {rec.confidence:.0%}")
        lines.append("")

        for f in schema.get("fields", []):
            fname = f["field_name"]
            dname = f.get("display_name", fname)
            val = rec.fields.get(fname, "N/A")

            if isinstance(val, list):
                lines.append(f"**{dname}**:")
                for item in val:
                    lines.append(f"  - {item}")
            else:
                lines.append(f"**{dname}**: {val}")

        lines.append("")
        lines.append("---")
        lines.append("")

    content = "\n".join(lines)
    output_path.write_text(content, encoding="utf-8")
    logger.info("markdown_exported", path=str(output_path))


def export_json(
    result: DocumentExtractionResult,
    output_path: str | Path,
    *,
    mask_phi: bool = False,
) -> None:
    """Export extraction results to JSON.

    Args:
        mask_phi: If True, redact PHI fields/values before serialisation.
    """
    output_path = Path(output_path)
    result = _apply_mask_phi(result, mask_phi=mask_phi)
    data = result.to_dict()
    output_path.write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info("json_exported", path=str(output_path), mask_phi=mask_phi)


def export_fhir_bundle(
    result: DocumentExtractionResult,
    output_path: str | Path,
    *,
    mask_phi: bool = False,
    profile: str | None = None,
) -> dict[str, Any] | None:
    """Phase K — emit a FHIR R4 Bundle for Healthcare-mode extractions.

    Gated by profile: only runs when ``profile == "medical-rcm"`` (or
    when the caller explicitly forces it by passing a known medical
    document_type via the underlying record). For non-medical profiles
    this is a no-op so the surrounding callsite stays profile-agnostic.

    The bundle is written to ``output_path`` as ``application/fhir+json``
    and also returned to the caller (so the API can stream it inline up
    to a size cap without re-reading the file).

    Args:
        result: Multi-record extraction result. The first record is
            treated as the primary subject; additional records are
            currently dropped (single-Patient bundle). A future
            ``profile_overlays`` extension can fan multi-record to
            one bundle per record.
        output_path: Path to write the FHIR Bundle JSON.
        mask_phi: If True, redact PHI fields/values before bundle
            construction.
        profile: Phase K profile id. When supplied and not equal to
            ``"medical-rcm"``, the call short-circuits and returns
            ``None``. ``None`` (auto-detected) defers the decision to
            the underlying record's ``document_type``.

    Returns:
        The bundle dict on success; ``None`` when emission is skipped
        (non-medical profile, no records, missing required fields).
    """
    from src.export.fhir_exporter import export_fhir

    output_path = Path(output_path)

    # Phase K — profile gate. ``None`` means caller did not pass a
    # profile, so we fall back to inferring from document_type below.
    if profile is not None and profile != "medical-rcm":
        logger.debug(
            "fhir_emission_skipped_non_medical_profile",
            profile=profile,
        )
        return None

    masked = _apply_mask_phi(result, mask_phi=mask_phi)
    if not masked.records:
        logger.info("fhir_emission_skipped_no_records", path=str(output_path))
        return None

    primary = masked.records[0]
    document_type = (masked.document_type or "").lower().strip()
    # DocumentExtractionResult doesn't carry a processing_id today (it's
    # threaded through state), so derive a stable bundle id from the pdf
    # path. The FHIR exporter falls back to a UUID when None.
    processing_id = getattr(masked, "processing_id", None) or Path(masked.pdf_path).stem

    # Profile=auto path: only emit FHIR when the document_type is a
    # known medical schema. This protects the General-mode path from
    # silently emitting FHIR for a contract or invoice.
    # Canonical + Gemma-analyzer-friendly aliases. The analyzer's
    # adaptive schema names land here verbatim, so we accept the
    # human-readable forms it produces as well as the short codes.
    medical_doc_types = {
        "cms1500",
        "ub04",
        "eob",
        "superbill",
        "health_insurance_claim_form",
        "hcfa_1500",
        "uniform_billing_04",
        "explanation_of_benefits",
        "remittance_advice",
    }
    if profile is None and document_type not in medical_doc_types:
        logger.debug(
            "fhir_emission_skipped_non_medical_doc_type",
            document_type=document_type,
        )
        return None

    fhir_bundle = export_fhir(
        record=primary.fields,
        document_type=document_type,
        processing_id=processing_id,
    )

    output_path.write_text(
        json.dumps(fhir_bundle.bundle, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    logger.info(
        "fhir_exported",
        path=str(output_path),
        document_type=document_type,
        resource_count=fhir_bundle.resource_count,
        validated=fhir_bundle.validated,
        mask_phi=mask_phi,
    )
    return fhir_bundle.bundle


def write_signed_receipt(
    *,
    bundle_dir: Path | str,
    processing_id: str,
    profile: str | None,
    artefact_paths: list[Path | str],
    audit_chain_tail: str | None = None,
    signing_key: str | None = None,
    signer_key_id: str | None = None,
) -> Path:
    """Phase K — mint a SignedReceipt over the export bundle.

    Hashes every artefact, optionally HMAC-signs the receipt, and writes
    ``receipt.json`` into ``bundle_dir``. The receipt is offline-verifiable
    with :func:`src.export.signed_receipt.verify_receipt`.

    Args:
        bundle_dir: Where ``receipt.json`` will land.
        processing_id: Pipeline processing id — binds the receipt to a
            specific extraction run.
        profile: Active mode profile (``"medical-rcm"`` /
            ``"generic-document"`` / etc.). Stored verbatim.
        artefact_paths: All files the receipt should attest. Missing
            files are silently skipped (Phase K: FHIR bundle is
            absent for non-medical extractions).
        audit_chain_tail: Optional audit-chain tail hash (Phase 8). When
            ``None`` the receipt records ``audit_chain_tail=None`` —
            still useful for offline artefact-integrity checks.
        signing_key: HMAC shared secret. ``None`` → unsigned receipt.
        signer_key_id: Optional key identifier.

    Returns:
        The path to the written ``receipt.json``.
    """
    from src.export.signed_receipt import mint_receipt, write_receipt

    bundle_dir_p = Path(bundle_dir)
    receipt = mint_receipt(
        processing_id=processing_id,
        profile=profile,
        artefact_paths=[Path(p) for p in artefact_paths],
        audit_chain_tail=audit_chain_tail,
        signing_key=signing_key,
        signer_key_id=signer_key_id,
    )
    out = write_receipt(receipt, bundle_dir_p / "receipt.json")
    logger.info(
        "signed_receipt_written",
        path=str(out),
        artefact_count=len(receipt.artefact_hashes),
        signed=receipt.signature is not None,
        signer_key_id=signer_key_id,
        profile=profile,
    )
    return out
