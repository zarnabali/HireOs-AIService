"""
Spreadsheet file processor for XLSX and CSV files.

Renders spreadsheet data as page images with table formatting,
enabling VLM-based extraction of tabular data.
"""

import csv
import io
import secrets
import time
from datetime import UTC, datetime
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont

from src.config import get_logger
from src.preprocessing.base_processor import (
    BaseFileProcessor,
    FileProcessingError,
    FileValidationError,
)
from src.preprocessing.pdf_processor import (
    PageImage,
    PDFMetadata,
    ProcessingResult,
)


logger = get_logger(__name__)

# Rendering constants
PAGE_WIDTH_PX = 3400    # Wide format for tables
PAGE_HEIGHT_PX = 2550   # Landscape orientation
MARGIN_PX = 100
ROW_HEIGHT = 40
HEADER_HEIGHT = 50
FONT_SIZE = 24
HEADER_FONT_SIZE = 28
MAX_ROWS_PER_PAGE = 50
MAX_FILE_SIZE_MB = 50
CELL_PADDING = 10


class SpreadsheetProcessor(BaseFileProcessor):
    """
    Process XLSX and CSV files into PageImage objects.

    Reads spreadsheet data and renders it as table images suitable
    for VLM extraction. Large sheets are split across multiple pages.
    """

    def __init__(self, dpi: int = 300) -> None:
        self._dpi = dpi

    def validate(self, file_path: Path) -> None:
        """Validate spreadsheet file."""
        if not file_path.exists():
            raise FileValidationError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            raise FileValidationError(f"Not a spreadsheet: {suffix}")

        size_mb = file_path.stat().st_size / (1024 * 1024)
        if size_mb > MAX_FILE_SIZE_MB:
            raise FileValidationError(
                f"File size {size_mb:.1f} MB exceeds limit of {MAX_FILE_SIZE_MB} MB"
            )

    def process(self, file_path: Path) -> ProcessingResult:
        """Convert spreadsheet to page images."""
        start_time = time.monotonic()
        self.validate(file_path)

        file_hash = self._compute_file_hash(file_path)
        processing_id = f"sheet_{secrets.token_hex(8)}"
        warnings: list[str] = []

        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            sheets = self._read_csv(file_path)
        else:
            sheets = self._read_xlsx(file_path, warnings)

        # Render each sheet into pages
        pages: list[PageImage] = []
        page_number = 1

        for sheet_name, rows in sheets:
            sheet_pages = self._render_sheet(rows, sheet_name, page_number)
            pages.extend(sheet_pages)
            page_number += len(sheet_pages)

        if not pages:
            warnings.append("No data found in spreadsheet")

        processing_time_ms = int((time.monotonic() - start_time) * 1000)

        metadata = PDFMetadata(
            file_path=file_path,
            file_name=file_path.name,
            file_size_bytes=file_path.stat().st_size,
            file_hash=file_hash,
            page_count=len(pages),
            title=None,
            author=None,
            subject=None,
            keywords=None,
            creator="SpreadsheetProcessor",
            producer="openpyxl/csv",
            creation_date=datetime.now(UTC),
            modification_date=None,
            pdf_version="N/A",
            is_encrypted=False,
            has_forms=False,
            has_annotations=False,
            processing_id=processing_id,
        )

        logger.info(
            "spreadsheet_processing_complete",
            file=file_path.name,
            pages=len(pages),
            sheets=len(sheets),
            time_ms=processing_time_ms,
        )

        return ProcessingResult(
            metadata=metadata,
            pages=pages,
            processing_time_ms=processing_time_ms,
            warnings=warnings,
        )

    def _read_csv(self, file_path: Path) -> list[tuple[str, list[list[str]]]]:
        """Read CSV file into rows."""
        rows: list[list[str]] = []
        with file_path.open("r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
        return [("Sheet1", rows)]

    def _read_xlsx(
        self, file_path: Path, warnings: list[str],
    ) -> list[tuple[str, list[list[str]]]]:
        """Read XLSX file into sheets with rows."""
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise FileProcessingError(
                "openpyxl is required for XLSX processing. "
                "Install with: pip install openpyxl"
            ) from e

        sheets: list[tuple[str, list[list[str]]]] = []
        wb = load_workbook(str(file_path), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in str_row):
                    rows.append(str_row)
            if rows:
                sheets.append((sheet_name, rows))
            else:
                warnings.append(f"Sheet '{sheet_name}' is empty, skipping")

        wb.close()
        return sheets

    def _render_sheet(
        self,
        rows: list[list[str]],
        sheet_name: str,
        start_page: int,
    ) -> list[PageImage]:
        """Render a sheet's rows into page images."""
        pages: list[PageImage] = []

        if not rows:
            return pages

        # Split rows into page chunks (first row is header, repeated on each page)
        header = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else rows

        for chunk_idx in range(0, max(len(data_rows), 1), MAX_ROWS_PER_PAGE):
            chunk = data_rows[chunk_idx : chunk_idx + MAX_ROWS_PER_PAGE]
            page_rows = [header] + chunk if header else chunk

            page_img = self._render_table_page(
                page_rows,
                sheet_name=sheet_name,
                page_number=start_page + len(pages),
                is_header_row=bool(header),
            )
            pages.append(page_img)

        return pages

    def _render_table_page(
        self,
        rows: list[list[str]],
        sheet_name: str,
        page_number: int,
        is_header_row: bool,
    ) -> PageImage:
        """Render a single page of table data as an image."""
        if not rows:
            rows = [[""]]

        num_cols = max(len(row) for row in rows)
        col_width = max(80, (PAGE_WIDTH_PX - 2 * MARGIN_PX) // max(num_cols, 1))

        total_height = MARGIN_PX * 2 + HEADER_HEIGHT + len(rows) * ROW_HEIGHT + 60
        page_height = max(PAGE_HEIGHT_PX, total_height)

        img = Image.new("RGB", (PAGE_WIDTH_PX, page_height), "white")
        draw = ImageDraw.Draw(img)

        try:
            font = ImageFont.truetype("arial.ttf", FONT_SIZE)
            header_font = ImageFont.truetype("arial.ttf", HEADER_FONT_SIZE)
        except OSError:
            font = ImageFont.load_default()
            header_font = font

        # Draw sheet name header
        y = MARGIN_PX
        draw.text((MARGIN_PX, y), f"Sheet: {sheet_name}", fill="black", font=header_font)
        y += HEADER_HEIGHT

        text_content_parts: list[str] = [f"Sheet: {sheet_name}"]

        # Draw table
        for row_idx, row in enumerate(rows):
            x = MARGIN_PX

            # Header row background
            if row_idx == 0 and is_header_row:
                draw.rectangle(
                    [x, y, x + num_cols * col_width, y + ROW_HEIGHT],
                    fill="#E0E0E0",
                )

            row_texts: list[str] = []
            for col_idx in range(num_cols):
                cell_text = row[col_idx] if col_idx < len(row) else ""
                # Truncate long cell values for rendering
                display_text = cell_text[:30] + "..." if len(cell_text) > 30 else cell_text

                draw.text(
                    (x + CELL_PADDING, y + CELL_PADDING),
                    display_text,
                    fill="black",
                    font=font,
                )

                # Draw cell border
                draw.rectangle(
                    [x, y, x + col_width, y + ROW_HEIGHT],
                    outline="#999999",
                )
                row_texts.append(cell_text)
                x += col_width

            text_content_parts.append(" | ".join(row_texts))
            y += ROW_HEIGHT

        # Convert to PNG
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        png_bytes = buf.getvalue()

        orientation = self._detect_orientation(PAGE_WIDTH_PX, page_height)

        return PageImage(
            page_number=page_number,
            image_bytes=png_bytes,
            width=PAGE_WIDTH_PX,
            height=page_height,
            dpi=self._dpi,
            orientation=orientation,
            original_width_pts=PAGE_WIDTH_PX * 72.0 / self._dpi,
            original_height_pts=page_height * 72.0 / self._dpi,
            has_text=True,
            has_images=False,
            rotation=0,
            text_content="\n".join(text_content_parts),
        )
